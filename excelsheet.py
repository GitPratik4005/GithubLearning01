from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# =====================================================
# SHEET 1: MASTER PLAN (DAYS 1–10)
# =====================================================
ws1 = wb.active
ws1.title = "Master Plan"

ws1.append([
    "Day", "Hour Block", "Focus Area",
    "Exact Task", "Output / Proof",
    "Completed", "Status"
])

MASTER_TASKS = [
    # DAY 1
    ("Day 1","H1-2","Python Core","List vs Tuple vs Set vs Dict","Notes"),
    ("Day 1","H3","Functions","*args, **kwargs, lambda, exceptions","Code"),
    ("Day 1","H4","OOP","Classes & inheritance","Mini class"),
    ("Day 1","H5","Complexity","Big-O, dict vs list","Explanation"),
    ("Day 1","H6","Practice","Two Sum + LRU cache","Code"),

    # DAY 2
    ("Day 2","H1-2","NumPy","Arrays, shapes, broadcasting","Notebook"),
    ("Day 2","H3-4","Pandas","Filter, groupby, merge","Notebook"),
    ("Day 2","H5","SQL","SELECT, JOIN, GROUP BY","SQL"),
    ("Day 2","H6","Mini Project","CSV → Pandas → SQL","Script"),

    # DAY 3
    ("Day 3","H1","ML Basics","Supervised vs Unsupervised","Notes"),
    ("Day 3","H2-3","Models","Linear, Logistic, Tree, KNN","Code"),
    ("Day 3","H4","Sklearn","Pipelines & scaling","Notebook"),
    ("Day 3","H5","Metrics","Precision, Recall","Examples"),
    ("Day 3","H6","Project","Regression + Classification","Repo"),

    # DAY 4
    ("Day 4","H1","APIs","REST & HTTP methods","Notes"),
    ("Day 4","H2","FastAPI","Endpoints & validation","API"),
    ("Day 4","H3","Embeddings","Cosine similarity","Notes"),
    ("Day 4","H4","Vector DB","FAISS / ChromaDB","Demo"),
    ("Day 4","H5-6","RAG","Document search system","Working RAG"),

    # DAY 5
    ("Day 5","H1","Git","Init, commit, push, branch","Repo"),
    ("Day 5","H2","System Design","ML pipeline overview","Diagram"),
    ("Day 5","H3-5","Capstone","AI Document Search","Complete project"),
    ("Day 5","H6","Interview Prep","Explain project","Recording"),

    # DAY 6
    ("Day 6","H1","Python Internals","Memory model","Explanation"),
    ("Day 6","H2","Mutability","Default args, deepcopy","Examples"),
    ("Day 6","H3","Garbage Collection","Reference counting","Notes"),
    ("Day 6","H4","GIL","CPU vs IO","Explanation"),
    ("Day 6","H5","Gotchas","is vs ==, late binding","Examples"),
    ("Day 6","H6","Verbal","Explain internals","Recording"),

    # DAY 7
    ("Day 7","H1","SQL","Execution order","Explanation"),
    ("Day 7","H2","Indexing","B-tree, composite index","Notes"),
    ("Day 7","H3","Optimization","EXPLAIN ANALYZE","SQL"),
    ("Day 7","H4","Pandas Perf","Vectorization","Notebook"),
    ("Day 7","H5","Integration","Python + SQL","Code"),
    ("Day 7","H6","Practice","Optimize queries","Before/After"),

    # DAY 8
    ("Day 8","H1","ML Theory","Bias vs variance","Explanation"),
    ("Day 8","H2","Feature Eng","Scaling, encoding","Examples"),
    ("Day 8","H3","Model Choice","Linear vs Tree","Notes"),
    ("Day 8","H4","Metrics","ROC-AUC","Examples"),
    ("Day 8","H5","Cross Validation","K-Fold & pipelines","Notebook"),
    ("Day 8","H6","Verbal","Explain improvements","Recording"),

    # DAY 9
    ("Day 9","H1","Embeddings","Dense vectors","Notes"),
    ("Day 9","H2","Vector Index","HNSW, IVF","Comparison"),
    ("Day 9","H3","RAG Failures","Hallucinations","Notes"),
    ("Day 9","H4","LLM APIs","Prompt, temperature","Examples"),
    ("Day 9","H5","Scaling RAG","Caching, reranking","Design"),
    ("Day 9","H6","Verbal","Vector DB vs SQL","Recording"),

    # DAY 10
    ("Day 10","H1","System Design","End-to-end ML system","Diagram"),
    ("Day 10","H2","Monitoring","Data & concept drift","Notes"),
    ("Day 10","H3","API Design","Versioning & auth","Design"),
    ("Day 10","H4","Git Advanced","Rebase vs merge","Git log"),
    ("Day 10","H5","Mock Interview","Full mock","Recording"),
    ("Day 10","H6","Resume","Final resume polish","Resume"),
]

for t in MASTER_TASKS:
    ws1.append(list(t) + ["", "Not Started"])

checkbox = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
status_dd = DataValidation(
    type="list",
    formula1='"Not Started,In Progress,Done,Explained Aloud"',
    allow_blank=True
)

ws1.add_data_validation(checkbox)
ws1.add_data_validation(status_dd)

for r in range(2, ws1.max_row + 1):
    checkbox.add(ws1[f"F{r}"])
    status_dd.add(ws1[f"G{r}"])

# =====================================================
# SHEET 2: DAILY SCORE
# =====================================================
ws2 = wb.create_sheet("Daily Score")
ws2.append([
    "Day","Planned Hours","Actual Hours",
    "Tasks Completed %","Confidence (1-5)",
    "Weak Areas","Next-Day Fix"
])
for i in range(1, 11):
    ws2.append([f"Day {i}","","","","","",""])

# =====================================================
# SHEET 3: INTERVIEW READINESS
# =====================================================
ws3 = wb.create_sheet("Interview Readiness")
ws3.append(["Skill Area","Explain Clearly","Code Live","Design Discussion"])
skills = [
    "Python Core","Python Internals","SQL & Indexing","Pandas Performance",
    "ML Fundamentals","ML Metrics","Vector Databases","RAG Systems",
    "API Design","Git","System Design"
]
for s in skills:
    ws3.append([s,"","",""])

yesno = DataValidation(type="list", formula1='"YES,NO"', allow_blank=True)
ws3.add_data_validation(yesno)
for col in ["B","C","D"]:
    for r in range(2, ws3.max_row + 1):
        yesno.add(ws3[f"{col}{r}"])

# =====================================================
# SHEET 4: 15-MIN REVISION
# =====================================================
ws4 = wb.create_sheet("15-Min Revision")
ws4.append([
    "Day","Topic","Key Question",
    "Your Answer","Confidence (1-5)","Revised"
])

revision = [
    ("Day 1","Python","Why is dict lookup O(1)?"),
    ("Day 2","Pandas","Why is apply() slow?"),
    ("Day 3","ML","What causes overfitting?"),
    ("Day 4","APIs","PUT vs POST?"),
    ("Day 5","System Design","ML pipeline steps"),
    ("Day 6","Python GIL","Why threads fail for CPU?"),
    ("Day 7","SQL","When index is ignored?"),
    ("Day 8","ML Metrics","Why accuracy misleads?"),
    ("Day 9","RAG","Top RAG failure modes"),
    ("Day 10","Monitoring","What is data drift?")
]

for r in revision:
    ws4.append(list(r) + ["",""])

for r in range(2, ws4.max_row + 1):
    checkbox.add(ws4[f"F{r}"])

# =====================================================
# SHEET 5: STATUS LOGIC (REFERENCE)
# =====================================================
ws5 = wb.create_sheet("Status Logic")
ws5.append(["Status","Meaning"])
ws5.append(["Not Started","Task not begun"])
ws5.append(["In Progress","Work ongoing"])
ws5.append(["Done","Completed"])
ws5.append(["Explained Aloud","Can explain without notes"])

# =====================================================
# SHEET 6: MOCK INTERVIEW MATRIX
# =====================================================
ws6 = wb.create_sheet("Mock Interview Matrix")
ws6.append(["Skill","Explain","Code","Design","Verdict"])
for s in skills:
    ws6.append([s,"","","",""])

ws6.add_data_validation(yesno)
for col in ["B","C","D","E"]:
    for r in range(2, ws6.max_row + 1):
        yesno.add(ws6[f"{col}{r}"])

# =====================================================
# SHEET 7: EXECUTION SCORE
# =====================================================
ws7 = wb.create_sheet("Execution Score")
ws7.append([
    "Day","Planned Hours","Actual Hours",
    "Focus Quality (1-5)",
    "Outcome Quality (1-5)",
    "One Brutal Truth Learned"
])
for i in range(1, 11):
    ws7.append([f"Day {i}","","","","",""])

# =====================================================
# SAVE FILE
# =====================================================
wb.save("10-Day-Python-AI-Tracker.xlsx")
print("✅ 10-Day Python AI Tracker created successfully")
